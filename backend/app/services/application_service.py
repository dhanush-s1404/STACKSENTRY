import math
import uuid
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from io import BytesIO
from sqlalchemy import select, func, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.application import Application, ApplicationStatus, PreferredWorkMode
from app.models.applicant import ApplicantProfile
from app.models.job import Job
from app.models.user import User
from app.models.status_history import ApplicationStatusHistory
from app.utils.security import generate_tracking_number


class ApplicationService:
    """Service for managing job applications."""

    @staticmethod
    async def create_application(
        db: AsyncSession,
        applicant_id: uuid.UUID,
        job_id: uuid.UUID,
        data: dict,
    ) -> Application:
        """Create a new job application."""
        existing = await db.execute(
            select(Application).where(
                and_(
                    Application.applicant_id == applicant_id,
                    Application.job_id == job_id,
                )
            )
        )
        if existing.scalar_one_or_none():
            raise ValueError("You have already applied for this job.")

        job = await db.get(Job, job_id)
        if not job:
            raise ValueError("Job not found.")
        if not job.is_active:
            raise ValueError("This job is no longer active.")
        if job.application_deadline and job.application_deadline < datetime.now(timezone.utc):
            raise ValueError("Application deadline has passed.")
        if job.positions_filled >= job.positions_available:
            raise ValueError("No positions available for this job.")

        tracking = generate_tracking_number()

        application = Application(
            tracking_number=tracking,
            applicant_id=applicant_id,
            job_id=job_id,
            status=ApplicationStatus.applied,
            cover_letter=data.get("cover_letter"),
            expected_salary=data.get("expected_salary"),
            preferred_work_mode=(
                PreferredWorkMode(data["preferred_work_mode"])
                if data.get("preferred_work_mode")
                else None
            ),
        )
        db.add(application)
        await db.flush()

        status_history = ApplicationStatusHistory(
            application_id=application.id,
            old_status=None,
            new_status=ApplicationStatus.applied.value,
            changed_by=applicant_id,
            notes="Application submitted",
        )
        db.add(status_history)

        job.positions_filled += 1

        await db.flush()
        return application

    @staticmethod
    async def get_application(
        db: AsyncSession, application_id: uuid.UUID
    ) -> Optional[Application]:
        """Get application by ID with relationships."""
        result = await db.execute(
            select(Application)
            .options(
                selectinload(Application.applicant).selectinload(ApplicantProfile.user),
                selectinload(Application.job),
                selectinload(Application.status_history),
            )
            .where(Application.id == application_id)
        )
        return result.scalar_one_or_none()

    @staticmethod
    async def list_applications(
        db: AsyncSession,
        filters: dict,
        page: int = 1,
        per_page: int = 20,
        applicant_id: Optional[uuid.UUID] = None,
    ) -> tuple[list[Application], int]:
        """List applications with filters and pagination."""
        query = select(Application).options(
            selectinload(Application.applicant).selectinload(ApplicantProfile.user),
            selectinload(Application.job),
        )

        if applicant_id:
            query = query.where(Application.applicant_id == applicant_id)

        if filters.get("status"):
            query = query.where(Application.status == filters["status"])

        if filters.get("job_id"):
            query = query.where(Application.job_id == filters["job_id"])

        if filters.get("search"):
            search = f"%{filters['search']}%"
            query = query.join(Application.applicant).join(ApplicantProfile.user).where(
                or_(
                    Application.tracking_number.ilike(search),
                    User.full_name.ilike(search),
                    User.email.ilike(search),
                )
            )

        if filters.get("date_from"):
            query = query.where(Application.created_at >= filters["date_from"])

        if filters.get("date_to"):
            query = query.where(Application.created_at <= filters["date_to"])

        count_query = select(func.count()).select_from(query.subquery())
        total = (await db.execute(count_query)).scalar() or 0

        query = query.order_by(Application.created_at.desc())
        query = query.offset((page - 1) * per_page).limit(per_page)

        result = await db.execute(query)
        applications = list(result.scalars().all())

        return applications, total

    @staticmethod
    async def update_status(
        db: AsyncSession,
        application_id: uuid.UUID,
        new_status: str,
        changed_by: uuid.UUID,
        notes: Optional[str] = None,
    ) -> Application:
        """Update application status with history tracking."""
        application = await db.get(Application, application_id)
        if not application:
            raise ValueError("Application not found.")

        old_status = application.status.value
        application.status = ApplicationStatus(new_status)
        application.status_updated_at = datetime.now(timezone.utc)

        if new_status == "rejected" and notes:
            application.rejection_reason = notes
        elif new_status in ("interview_scheduled", "interview_completed", "shortlisted"):
            application.hr_notes = notes

        status_history = ApplicationStatusHistory(
            application_id=application_id,
            old_status=old_status,
            new_status=new_status,
            changed_by=changed_by,
            notes=notes,
        )
        db.add(status_history)
        await db.flush()
        return application

    @staticmethod
    async def get_dashboard_stats(db: AsyncSession) -> dict:
        """Get dashboard statistics."""
        total_applications = (
            await db.execute(select(func.count(Application.id)))
        ).scalar() or 0

        status_counts = {}
        for status in ApplicationStatus:
            count = (
                await db.execute(
                    select(func.count(Application.id)).where(
                        Application.status == status
                    )
                )
            ).scalar() or 0
            status_counts[status.value] = count

        active_jobs = (
            await db.execute(
                select(func.count(Job.id)).where(Job.is_active == True)
            )
        ).scalar() or 0

        recent_applications = (
            await db.execute(
                select(func.count(Application.id)).where(
                    Application.created_at
                    >= datetime.now(timezone.utc).replace(
                        hour=0, minute=0, second=0, microsecond=0
                    )
                )
            )
        ).scalar() or 0

        return {
            "total_applications": total_applications,
            "status_counts": status_counts,
            "active_jobs": active_jobs,
            "recent_applications": recent_applications,
        }

    @staticmethod
    def calculate_resume_score(
        applicant_data: dict, job_requirements: str
    ) -> float:
        """Calculate a basic resume score based on applicant data and job requirements."""
        score = 0.0
        max_score = 100.0

        if applicant_data.get("experience_years", 0) > 0:
            score += min(applicant_data["experience_years"] * 5, 25)

        if applicant_data.get("cgpa"):
            score += min(applicant_data["cgpa"] * 5, 20)

        if applicant_data.get("skills"):
            score += min(len(applicant_data["skills"]) * 3, 25)

        if applicant_data.get("projects"):
            score += min(len(applicant_data["projects"]) * 5, 15)

        if applicant_data.get("github_url") or applicant_data.get("linkedin_url"):
            score += 10

        if applicant_data.get("bio"):
            score += 5

        return min(score, max_score)

    @staticmethod
    def export_to_excel(applications: list[Application]) -> BytesIO:
        """Export applications to Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"

        headers = [
            "Tracking Number", "Applicant Name", "Email", "Job Title",
            "Status", "Applied Date", "Resume Score", "Expected Salary",
        ]
        ws.append(headers)

        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(
            start_color="2563EB", end_color="2563EB", fill_type="solid"
        )
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for app in applications:
            ws.append([
                app.tracking_number,
                app.applicant.user.full_name if app.applicant and app.applicant.user else "N/A",
                app.applicant.user.email if app.applicant and app.applicant.user else "N/A",
                app.job.title if app.job else "N/A",
                app.status.value,
                app.created_at.strftime("%Y-%m-%d %H:%M"),
                app.resume_score or "N/A",
                app.expected_salary or "N/A",
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_to_pdf(applications: list[Application]) -> BytesIO:
        """Export applications to PDF file."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Title"], fontSize=18,
            textColor=colors.HexColor("#2563EB"), spaceAfter=20,
        )
        elements.append(Paragraph("StackSentry Technologies - Applications Report", title_style))
        elements.append(Paragraph(
            f"Generated on: {datetime.now(timezone.utc).strftime('%B %d, %Y at %I:%M %p UTC')}",
            styles["Normal"],
        ))
        elements.append(Spacer(1, 20))

        data = [["#", "Tracking No.", "Applicant", "Job Title", "Status", "Date"]]
        for idx, app in enumerate(applications, 1):
            data.append([
                str(idx),
                app.tracking_number,
                app.applicant.user.full_name if app.applicant and app.applicant.user else "N/A",
                app.job.title if app.job else "N/A",
                app.status.value.replace("_", " ").title(),
                app.created_at.strftime("%Y-%m-%d"),
            ])

        table = Table(data, colWidths=[0.4 * inch, 1.2 * inch, 1.5 * inch, 1.5 * inch, 1.2 * inch, 1 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 1), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)

        doc.build(elements)
        output.seek(0)
        return output
